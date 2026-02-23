"""
Module for generating reports from evaluation outputs.
"""
import asyncio
import json
import logging
import shutil
from pathlib import Path
from typing import List, Dict

import pandas as pd
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


class ExcelReportGenerator:
    """Generates Excel reports from evaluation outputs."""

    async def generate(
        self,
        config_path: str,
        outputs_dir: str,
        eval_dir: str,
        template_path: str | None
    ) -> None:
        """
        Generates an Excel report from JSON output files.

        Args:
            config_path: Path to the configuration file used for the evaluation.
            outputs_dir: Directory containing the JSON output files.
            eval_dir: Directory where the evaluation report should be saved.
            template_path: Path to the Excel report template.
        """
        if not template_path:
            logger.warning(
                "No Excel report template specified. Skipping report generation.")
            return

        config_path_obj = Path(config_path)
        output_path_obj = Path(outputs_dir)
        eval_path_obj = Path(eval_dir)
        template_path_obj = Path(template_path)

        if not template_path_obj.exists():
            logger.error(f"Excel report template not found: {template_path}")
            return

        # Determine report filename
        report_filename = f"{config_path_obj.stem}-report.xlsx"
        report_path = eval_path_obj / report_filename

        # Copy template to report path
        try:
            await asyncio.to_thread(
                shutil.copy2, template_path_obj, report_path
            )
            logger.info(f"Created report file: {report_path}")
        except Exception as e:
            logger.error(f"Failed to copy report template: {e}")
            return

        # Enumerate JSON files
        json_files = sorted(output_path_obj.glob("*.json"))
        if not json_files:
            logger.warning(
                f"No JSON output files found in {outputs_dir}. Report will be empty.")
            return

        data: List[Dict] = []

        # Define column order
        columns = [
            "reference_file",
            "model_name",
            "model_label",
            "wer",
            "mer",
            "wil",
            "cer",
            "normalized_wer",
            "normalized_mer",
            "normalized_wil",
            "normalized_cer",
            "audio_file_name",
            "audio_file_duration"
        ]

        # Read all JSON files
        for json_file in json_files:
            try:
                item = await asyncio.to_thread(self._read_json_file, json_file)
                if item:
                    row = {col: item.get(col) for col in columns}
                    data.append(row)
            except json.JSONDecodeError:
                logger.error(f"Failed to parse JSON file: {json_file}")
            except Exception as e:
                logger.error(f"Error reading {json_file}: {e}")

        if not data:
            return

        df = pd.DataFrame(data, columns=columns)

        try:
            # Write Excel report
            await asyncio.to_thread(
                self._write_excel_report,
                report_path,
                df
            )
            logger.info(
                f"Successfully added {len(data)} rows to {report_path}")

        except Exception as e:
            logger.error(f"Failed to write to Excel report: {e}")

    @staticmethod
    def _read_json_file(json_file: Path) -> dict | None:
        """Read and return JSON file contents."""
        with open(json_file, 'r', encoding='utf-8') as f:
            return json.load(f)

    @staticmethod
    def _write_excel_report(report_path: Path, df: pd.DataFrame) -> None:
        """Write DataFrame to Excel report."""
        # Load workbook to find the first empty row
        # We need to verify the sheet exists
        wb = load_workbook(report_path)
        sheet_name = 'Outputs'

        if sheet_name not in wb.sheetnames:
            logger.error(
                f"Sheet '{sheet_name}' not found in template. Cannot append data.")
            return

        ws = wb[sheet_name]
        start_row = ws.max_row
        wb.close()  # Close to allow pandas to write

        # Append data using pandas
        with pd.ExcelWriter(
            report_path,
            engine='openpyxl',
            mode='a',
            if_sheet_exists='overlay'
        ) as writer:
            df.to_excel(
                writer,
                sheet_name=sheet_name,
                startrow=start_row,
                header=False,
                index=False
            )


async def generate_excel_report(
    config_path: str,
    outputs_dir: str,
    eval_dir: str,
    template_path: str | None
) -> None:
    """
    Generates an Excel report from JSON output files.

    Args:
        config_path: Path to the configuration file used for the evaluation.
        outputs_dir: Directory containing the JSON output files.
        eval_dir: Directory where the evaluation report should be saved.
        template_path: Path to the Excel report template.
    """
    generator = ExcelReportGenerator()
    await generator.generate(config_path, outputs_dir, eval_dir, template_path)
